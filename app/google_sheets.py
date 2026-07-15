"""Export generated Excel reports to a configured Google spreadsheet."""

from __future__ import annotations

import json
import os
import re
import secrets
import threading
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SHEETS_SCOPE = "https://www.googleapis.com/auth/spreadsheets"
MAX_SHEET_TITLE = 100
VALUE_RANGE_ROWS = 500
MAX_VALUE_BATCH_BYTES = 1_800_000
FORMAT_CHUNK_REQUESTS = 450
RETRYABLE_STATUSES = {429, 500, 502, 503, 504}
MAX_API_ATTEMPTS = 5
INVALID_TITLE_CHARS = re.compile(r"[\\/?*\[\]:]")
GOOGLE_EXPORT_LOCK = threading.Lock()


def credentials_path() -> Path:
    """Return the configured service-account file without exposing its contents."""
    configured = _configured_value(
        "GOOGLE_CREDENTIALS_FILE",
        "Не задана переменная GOOGLE_CREDENTIALS_FILE",
    )
    path = Path(configured)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    if not path.is_file():
        raise RuntimeError(f"Файл Google credentials не найден: {path}")
    return path


def export_workbook(path: str | Path, project: str, analysis_date: date) -> dict[str, Any]:
    """Append every workbook sheet as a new tab in the configured spreadsheet."""
    with GOOGLE_EXPORT_LOCK:
        return _export_workbook(path, project, analysis_date)


def _export_workbook(path: str | Path, project: str, analysis_date: date) -> dict[str, Any]:
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    spreadsheet_id = _configured_value(
        "GOOGLE_SPREADSHEET_ID",
        "Не задана переменная GOOGLE_SPREADSHEET_ID",
    )
    credentials = Credentials.from_service_account_file(
        credentials_path(),
        scopes=[SHEETS_SCOPE],
    )
    service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    metadata = _execute(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="properties(title),sheets(properties(sheetId,title))",
        )
    )
    existing_titles = {
        str(sheet["properties"]["title"])
        for sheet in metadata.get("sheets", [])
    }

    existing_sheet_ids = {
        int(sheet["properties"]["sheetId"])
        for sheet in metadata.get("sheets", [])
    }
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    created: list[dict[str, Any]] = []
    try:
        plans: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            title = unique_sheet_title(
                project=project,
                analysis_date=analysis_date,
                source_title=worksheet.title,
                existing_titles=existing_titles,
            )
            existing_titles.add(title)
            sheet_id = _new_sheet_id(existing_sheet_ids)
            existing_sheet_ids.add(sheet_id)
            fill_runs = _cell_fill_runs(worksheet)
            headers = [
                str(cell.value or "")
                for cell in next(worksheet.iter_rows(min_row=1, max_row=1), [])
            ]
            plans.append(
                {
                    "worksheet": worksheet,
                    "title": title,
                    "sheet_id": sheet_id,
                    "headers": headers,
                    "fill_runs": fill_runs,
                }
            )
            created.append({"title": title, "sheet_id": str(sheet_id)})
        _create_and_format_sheets(service, spreadsheet_id, plans)
        _write_all_values(service, spreadsheet_id, plans)
    finally:
        workbook.close()

    first_gid = created[0]["sheet_id"] if created else "0"
    return {
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_title": str(metadata.get("properties", {}).get("title", "")),
        "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={first_gid}",
        "sheets": created,
    }


def unique_sheet_title(
    project: str,
    analysis_date: date,
    source_title: str,
    existing_titles: set[str],
) -> str:
    """Build a valid, readable, collision-safe Google sheet title."""
    parts = [project.strip(), analysis_date.isoformat(), source_title.strip()]
    base = INVALID_TITLE_CHARS.sub("-", " — ".join(part for part in parts if part)).strip(" '")
    base = base or "Анализ"
    base = base[:MAX_SHEET_TITLE].rstrip()
    if base not in existing_titles:
        return base
    counter = 2
    while True:
        suffix = f" ({counter})"
        candidate = f"{base[: MAX_SHEET_TITLE - len(suffix)].rstrip()}{suffix}"
        if candidate not in existing_titles:
            return candidate
        counter += 1


def google_value(value: Any) -> str | int | float | bool | None:
    """Convert an openpyxl cell value to a Sheets API-compatible scalar."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _configured_value(name: str, missing_message: str) -> str:
    load_dotenv(PROJECT_ROOT / ".env")
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(missing_message)
    return value


def _new_sheet_id(existing_ids: set[int]) -> int:
    while True:
        sheet_id = secrets.randbelow(2_000_000_000) + 1
        if sheet_id not in existing_ids:
            return sheet_id


def _create_and_format_sheets(
    service: Any,
    spreadsheet_id: str,
    plans: list[dict[str, Any]],
) -> None:
    requests: list[dict[str, Any]] = []
    for plan in plans:
        worksheet = plan["worksheet"]
        requests.append(
            {
                "addSheet": {
                    "properties": {
                        "sheetId": plan["sheet_id"],
                        "title": plan["title"],
                        "gridProperties": {
                            "rowCount": max(worksheet.max_row, 1),
                            "columnCount": max(worksheet.max_column, 1),
                        },
                    }
                }
            }
        )
    for plan in plans:
        worksheet = plan["worksheet"]
        requests.extend(
            _format_requests(
                plan["sheet_id"],
                worksheet.max_row,
                worksheet.max_column,
                plan["headers"],
                plan["fill_runs"],
            )
        )
    for start in range(0, len(requests), FORMAT_CHUNK_REQUESTS):
        _execute(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests[start : start + FORMAT_CHUNK_REQUESTS]},
            )
        )


def _write_all_values(
    service: Any,
    spreadsheet_id: str,
    plans: list[dict[str, Any]],
) -> None:
    batch: list[dict[str, Any]] = []
    batch_bytes = 0
    for value_range in _iter_value_ranges(plans):
        range_bytes = len(json.dumps(value_range, ensure_ascii=False).encode("utf-8"))
        if batch and batch_bytes + range_bytes > MAX_VALUE_BATCH_BYTES:
            _write_value_batch(service, spreadsheet_id, batch)
            batch = []
            batch_bytes = 0
        batch.append(value_range)
        batch_bytes += range_bytes
    if batch:
        _write_value_batch(service, spreadsheet_id, batch)


def _iter_value_ranges(plans: list[dict[str, Any]]) -> Iterator[dict[str, Any]]:
    for plan in plans:
        escaped_title = str(plan["title"]).replace("'", "''")
        chunk: list[list[Any]] = []
        start_row = 1
        for row in plan["worksheet"].iter_rows():
            chunk.append([google_value(cell.value) for cell in row])
            if len(chunk) >= VALUE_RANGE_ROWS:
                yield {"range": f"'{escaped_title}'!A{start_row}", "values": chunk}
                start_row += len(chunk)
                chunk = []
        if chunk:
            yield {"range": f"'{escaped_title}'!A{start_row}", "values": chunk}


def _write_value_batch(
    service: Any,
    spreadsheet_id: str,
    data: list[dict[str, Any]],
) -> None:
    _execute(
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"valueInputOption": "RAW", "data": data},
        )
    )


def _execute(request: Any) -> Any:
    for attempt in range(MAX_API_ATTEMPTS):
        try:
            return request.execute()
        except HttpError as exc:
            if exc.resp.status not in RETRYABLE_STATUSES or attempt == MAX_API_ATTEMPTS - 1:
                raise
            time.sleep(2**attempt)
    raise RuntimeError("Google API request retry loop ended unexpectedly")


def _format_requests(
    sheet_id: int,
    row_count: int,
    column_count: int,
    headers: list[str],
    fill_runs: list[tuple[int, int, int, str]],
) -> list[dict[str, Any]]:
    if row_count <= 0 or column_count <= 0:
        return []
    requests: list[dict[str, Any]] = [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": column_count,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColorStyle": {
                            "rgbColor": {"red": 0.9, "green": 0.92, "blue": 0.95}
                        },
                        "textFormat": {"bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColorStyle,textFormat)",
            }
        },
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "autoResizeDimensions": {
                "dimensions": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": column_count,
                }
            }
        },
    ]
    if row_count > 1:
        requests.append(
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": row_count,
                            "startColumnIndex": 0,
                            "endColumnIndex": column_count,
                        }
                    }
                }
            }
        )
    for index, header in enumerate(headers):
        if header.endswith("%") and row_count > 1:
            requests.append(
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "endRowIndex": row_count,
                            "startColumnIndex": index,
                            "endColumnIndex": index + 1,
                        },
                        "cell": {"userEnteredFormat": {"numberFormat": {"type": "PERCENT", "pattern": "0.00%"}}},
                        "fields": "userEnteredFormat.numberFormat",
                    }
                }
            )
    for column_index, start_row, end_row, color in fill_runs:
        requests.append(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_row,
                        "endRowIndex": end_row,
                        "startColumnIndex": column_index,
                        "endColumnIndex": column_index + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColorStyle": {"rgbColor": _rgb_color(color)}
                        }
                    },
                    "fields": "userEnteredFormat.backgroundColorStyle",
                }
            }
        )
    return requests


def _cell_fill_runs(worksheet: Any) -> list[tuple[int, int, int, str]]:
    """Return contiguous non-header fill ranges as zero-based coordinates."""
    active: dict[int, tuple[str, int]] = {}
    runs: list[tuple[int, int, int, str]] = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=1):
        seen_columns: set[int] = set()
        for column_index, cell in enumerate(row):
            color = _solid_fill_color(cell)
            if color is None:
                continue
            seen_columns.add(column_index)
            current = active.get(column_index)
            if current and current[0] == color:
                continue
            if current:
                runs.append((column_index, current[1], row_index, current[0]))
            active[column_index] = (color, row_index)
        for column_index in set(active) - seen_columns:
            color, start_row = active.pop(column_index)
            runs.append((column_index, start_row, row_index, color))
    end_row = max(worksheet.max_row, 1)
    for column_index, (color, start_row) in active.items():
        runs.append((column_index, start_row, end_row, color))
    return runs


def _solid_fill_color(cell: Any) -> str | None:
    fill = cell.fill
    if fill.fill_type != "solid" or fill.fgColor.type != "rgb" or not fill.fgColor.rgb:
        return None
    return str(fill.fgColor.rgb)[-6:].upper()


def _rgb_color(hex_color: str) -> dict[str, float]:
    return {
        "red": int(hex_color[0:2], 16) / 255,
        "green": int(hex_color[2:4], 16) / 255,
        "blue": int(hex_color[4:6], 16) / 255,
    }
