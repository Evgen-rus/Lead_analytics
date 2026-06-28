from openpyxl.styles import PatternFill


GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GRAY = PatternFill("solid", fgColor="D9EAD3")
BLUE = PatternFill("solid", fgColor="BDD7EE")
HEADER = PatternFill("solid", fgColor="D9EAF7")


def fill_for_metric(header: str, value: float):
    if value is None:
        return None
    v = float(value)
    if header in {"Кач. %", "Уже наши / купил %"}:
        if v == 0:
            return GRAY
        if v >= 0.03:
            return GREEN
        if v >= 0.015:
            return YELLOW
        return RED
    if header == "Рабочий потенциал %":
        if v == 0:
            return GRAY
        if v >= 0.05:
            return GREEN
        if v >= 0.02:
            return YELLOW
        return RED
    if header == "Сигнал спроса %":
        if v == 0:
            return GRAY
        if v >= 0.10:
            return GREEN
        if v >= 0.05:
            return YELLOW
        return RED
    if header == "Недозвон %":
        if v <= 0.15:
            return GREEN
        if v <= 0.25:
            return YELLOW
        return RED
    if header == "Некач. %":
        if v <= 0.60:
            return GREEN
        if v <= 0.75:
            return YELLOW
        return RED
    if header == "Не подходит по гео %":
        if v <= 0.05:
            return GREEN
        if v <= 0.15:
            return YELLOW
        return RED
    if header == "Требует проверки %":
        return GRAY if v == 0 else BLUE
    return None
