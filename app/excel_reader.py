from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def list_sheets(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    return wb.sheetnames


def read_excel_sheet(path: str | Path, sheet_name: str | None = None) -> pd.DataFrame:
    sheet = sheet_name if sheet_name is not None else 0
    df = pd.read_excel(path, sheet_name=sheet, dtype=object)
    df.columns = [str(col).strip() for col in df.columns]
    df = df.dropna(how="all")
    return df


def workbook_preview(path: str | Path, rows: int = 10) -> dict[str, pd.DataFrame]:
    return pd.read_excel(path, sheet_name=None, nrows=rows, dtype=object)
