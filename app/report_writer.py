from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font

from app.report_styles import HEADER, fill_for_metric


def write_excel(path: str | Path, sheets: dict[str, pd.DataFrame | list[str]]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, data in sheets.items():
            sheet_name = name[:31]
            if isinstance(data, list):
                df = pd.DataFrame({"Вывод": data})
            else:
                df = data
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            style_worksheet(writer.sheets[sheet_name], df)
    return path


def style_worksheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(bold=True)

    for idx, header in enumerate(df.columns, 1):
        values = df.iloc[:, idx - 1].dropna().astype(str)
        value_width = values.str.len().max() if not values.empty else 0
        max_len = min(max(len(str(header or "")), int(value_width)), 60)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max_len + 2

        if not str(header).endswith("%"):
            continue
        for column in ws.iter_cols(min_row=2, min_col=idx, max_row=ws.max_row, max_col=idx):
            for cell in column:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00%"
                    fill = fill_for_metric(str(header), cell.value)
                    if fill:
                        cell.fill = fill
