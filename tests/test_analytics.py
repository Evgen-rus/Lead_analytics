import pandas as pd
from openpyxl import load_workbook

from app import pipeline
from app.analytics import summarize
from app.models import ColumnMapping
from app.report_writer import write_excel


def test_summarize_excludes_do_not_count_group():
    df = pd.DataFrame(
        {
            "Канал": ["A", "A", "A"],
            "Группа статуса": ["Качественные", "Недозвон", "Не учитывать"],
        }
    )
    result = summarize(df, ["Канал"])
    assert result.loc[0, "Всего идентификаций"] == 2
    assert result.loc[0, "Сигнал спроса"] == 1
    assert result.loc[0, "Сигнал спроса %"] == 0.5


def test_summarize_places_demand_signal_after_already_bought_percent():
    df = pd.DataFrame(
        {
            "Группа статуса": ["Качественные", "Уже наши / уже купил", "Недозвон"],
        }
    )
    result = summarize(df, [])
    columns = list(result.columns)
    already_percent_index = columns.index("Уже наши / купил %")
    assert columns[already_percent_index + 1] == "Сигнал спроса"
    assert columns[already_percent_index + 2] == "Сигнал спроса %"
    assert columns[already_percent_index + 3] == "Недозвон"


def test_analyze_file_loads_status_rules_once(monkeypatch, tmp_path):
    source = pd.DataFrame(
        {
            "Дата": ["2026-01-01", "2026-01-02"],
            "Телефон": ["1", "2"],
            "Источник": ["example.com", "example.com"],
            "Статус": ["Новый", "Новый"],
        }
    )
    mapping = ColumnMapping(
        sheet_name="Сопоставленные",
        date_column="Дата",
        phone_column="Телефон",
        source_column="Источник",
        status_column="Статус",
    )
    rules = [object()]
    calls = 0

    monkeypatch.setattr(pipeline, "read_excel_sheet", lambda *_: source)

    def load_rules(project):
        nonlocal calls
        calls += 1
        return rules

    def classify_with_rules(status, comment, project, loaded_rules):
        assert loaded_rules is rules
        return "Качественные", "тест"

    monkeypatch.setattr(pipeline, "sorted_rules", load_rules)
    monkeypatch.setattr(pipeline, "classify", classify_with_rules)
    monkeypatch.setattr(pipeline, "write_excel", lambda path, _: path)

    pipeline.analyze_file("p", "input.xlsx", mapping, tmp_path)

    assert calls == 1


def test_write_excel_preserves_report_formatting(tmp_path):
    path = write_excel(
        tmp_path / "report.xlsx",
        {
            "Итог": pd.DataFrame(
                {"Кач. %": [0.03, 0.0910891089108911], "Канал": ["Поиск", "Сети"]}
            )
        },
    )

    ws = load_workbook(path).active

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:B3"
    assert ws["A2"].number_format == "0.00%"
    assert ws["A3"].number_format == "0.00%"
    assert ws["A2"].fill.fill_type == "solid"
    assert ws["A3"].fill.fill_type == "solid"
    assert ws.column_dimensions["B"].width >= len("Канал") + 2
