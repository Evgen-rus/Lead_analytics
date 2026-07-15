from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from app.google_sheets import (
    _cell_fill_runs,
    _create_and_format_sheets,
    _rgb_color,
    _write_all_values,
    google_value,
    unique_sheet_title,
)


class FakeRequest:
    def __init__(self, response: dict | None = None) -> None:
        self.response = response or {}

    def execute(self) -> dict:
        return self.response


class FakeSheetsService:
    def __init__(self) -> None:
        self.structure_batches: list[dict] = []
        self.value_batches: list[dict] = []

    def spreadsheets(self):
        return self

    def values(self):
        return self

    def batchUpdate(self, **kwargs):
        body = kwargs["body"]
        if "requests" in body:
            self.structure_batches.append(body)
        else:
            self.value_batches.append(body)
        return FakeRequest()


def test_unique_sheet_title_contains_project_date_and_source() -> None:
    title = unique_sheet_title("Проект 1", date(2026, 7, 15), "Итог", set())

    assert title == "Проект 1 — 2026-07-15 — Итог"


def test_unique_sheet_title_sanitizes_limits_and_deduplicates() -> None:
    original = unique_sheet_title("Проект/клиент", date(2026, 7, 15), "Данные: все", set())
    duplicate = unique_sheet_title(
        "Проект/клиент",
        date(2026, 7, 15),
        "Данные: все",
        {original},
    )
    long_title = unique_sheet_title("П" * 150, date(2026, 7, 15), "Итог", set())

    assert original == "Проект-клиент — 2026-07-15 — Данные- все"
    assert duplicate == f"{original} (2)"
    assert len(long_title) == 100


def test_google_value_keeps_scalars_and_converts_dates() -> None:
    assert google_value(None) is None
    assert google_value(42) == 42
    assert google_value(0.25) == 0.25
    assert google_value(True) is True
    assert google_value(date(2026, 7, 15)) == "2026-07-15"
    assert google_value(datetime(2026, 7, 15, 10, 30)) == "2026-07-15T10:30:00"


def test_cell_fill_runs_group_adjacent_colors_and_keep_changes() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    worksheet.append(["Кач. %", "Без цвета"])
    worksheet.append([0.03, "a"])
    worksheet.append([0.04, "b"])
    worksheet.append([0.01, "c"])
    worksheet["A2"].fill = green
    worksheet["A3"].fill = green
    worksheet["A4"].fill = red

    assert _cell_fill_runs(worksheet) == [
        (0, 1, 3, "C6EFCE"),
        (0, 3, 4, "FFC7CE"),
    ]


def test_rgb_color_converts_excel_hex_to_google_components() -> None:
    assert _rgb_color("FF8000") == {"red": 1.0, "green": 128 / 255, "blue": 0.0}


def test_structure_formatting_and_values_are_batched_across_sheets() -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Итог"
    first.append(["Кач. %"])
    first.append([0.03])
    second = workbook.create_sheet("Данные")
    second.append(["Статус"])
    second.append(["Качественный"])
    plans = [
        {
            "worksheet": first,
            "title": "Проект — 2026-07-15 — Итог",
            "sheet_id": 101,
            "headers": ["Кач. %"],
            "fill_runs": [],
        },
        {
            "worksheet": second,
            "title": "Проект — 2026-07-15 — Данные",
            "sheet_id": 102,
            "headers": ["Статус"],
            "fill_runs": [],
        },
    ]
    service = FakeSheetsService()

    _create_and_format_sheets(service, "spreadsheet", plans)
    _write_all_values(service, "spreadsheet", plans)

    assert len(service.structure_batches) == 1
    assert len(service.value_batches) == 1
    assert len(service.value_batches[0]["data"]) == 2
